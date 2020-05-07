import openpyxl

path = "./TC001.xlsx"

Work_Book = openpyxl.load_workbook(path)

Sheet = Work_Book.get_sheet_by_name("Sheet1")

rows = Sheet.max_row
cols = Sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(Sheet.cell(r,c).value, end=("       "))
    print()
