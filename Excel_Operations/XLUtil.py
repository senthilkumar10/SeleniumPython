import openpyxl

def getRowCount(xl_path, sheetname):
    Work_Book = openpyxl.load_workbook(xl_path)
    Sheet = Work_Book.get_sheet_by_name(sheetname)
    return (Sheet.max_row)



def getColumnCount(xl_path, sheetname):
    Work_Book = openpyxl.load_workbook(xl_path)
    Sheet = Work_Book.get_sheet_by_name(sheetname)
    return (Sheet.max_column)


def readData(xl_path, sheetname,rownum,colnum):
    Work_Book = openpyxl.load_workbook(xl_path)
    Sheet = Work_Book.get_sheet_by_name(sheetname)
    return (Sheet.cell(rownum,colnum).value)



def writeData(xl_path, sheetname,rownum,colnum,data):
    Work_Book = openpyxl.load_workbook(xl_path)
    Sheet = Work_Book.get_sheet_by_name(sheetname)
    Sheet.cell(rownum,colnum).value=data
    Work_Book.save(xl_path)

    